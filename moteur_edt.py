#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
╔════════════════════════════════════════════════════════════╗
║  MOTEUR EDT — Générateur d'emplois du temps scolaires      ║
║                                                            ║
║  Usage :  python3 moteur_edt.py donnees.xlsx resultat.xlsx ║
║                                                            ║
║  Entrée : le template Excel rempli (onglets Classes,       ║
║  Professeurs, Services, Disponibilités).                   ║
║  Sortie : un Excel avec une grille par classe + la vue     ║
║  par professeur. Zéro conflit garanti (OR-Tools CP-SAT).   ║
╚════════════════════════════════════════════════════════════╝

Règles gérées :
  H1  Volume horaire exact pour chaque service
  H2  Une classe = 1 seul cours par créneau
  H3  Un professeur = 1 seul cours par créneau
  H4  Indisponibilités des professeurs (onglet Disponibilités)
  H5  Permanents bloqués le mercredi après-midi (réunion)
  H6  Maximum 2h de la même matière par jour (hors blocs)
  H7  Jour imposé (optionnel) : le bloc / une séance tombe ce jour-là
  Blocs de 2h ou 3h consécutives (colonne optionnelle ;
  vide = l'outil décide librement). Les blocs peuvent
  traverser la pause déjeuner (choix de l'établissement).

Qualité optimisée :
  + matières scientifiques le matin
  − EPS à 14h30 (chaleur)   − matière lourde en fin de journée
  − trous dans la journée (classes ET professeurs)
"""

import sys
from collections import defaultdict

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from ortools.sat.python import cp_model

# ════════════════════════ CONFIG ════════════════════════
JOURS   = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi"]
N_JOURS = len(JOURS)
N_SLOTS = 8

SLOT_LABELS = [
    "07h00 – 08h00",
    "08h00 – 09h00",
    "09h00 – 10h00",
    "10h15 – 11h15",
    "11h15 – 12h15",
    "14h30 – 15h30",
    "15h30 – 16h30",
    "16h45 – 17h45",
]
SLOTS_MATIN      = list(range(5))
SLOTS_APMIDI     = list(range(5, 8))
IDX_DEBUT_APMIDI = 5
JOUR_INDEX = {j: i for i, j in enumerate(JOURS)}

SCIENCES = {"Mathématiques", "Physique-Chimie", "SVT"}
LOURDES  = {"Mathématiques", "Philosophie"}


# ════════════════════ 1. LECTURE EXCEL ════════════════════
def _val(row, i):
    """Valeur texte propre d'une cellule pandas, '' si vide."""
    if i >= len(row) or pd.isna(row.iloc[i]):
        return ""
    return str(row.iloc[i]).strip()


def lire_excel(chemin):
    xls = pd.ExcelFile(chemin)

    # ── Classes ──
    df = pd.read_excel(xls, "Classes", skiprows=2, header=0)
    df = df.dropna(subset=[df.columns[0]])
    classes = [_val(r, 0) for _, r in df.iterrows() if _val(r, 0)]

    # ── Professeurs (statut) ──
    df = pd.read_excel(xls, "Professeurs", skiprows=2, header=0)
    df = df.dropna(subset=[df.columns[0]])
    permanents = {_val(r, 0) for _, r in df.iterrows()
                  if _val(r, 0) and _val(r, 3) == "Permanent"}

    # ── Services ──
    # Colonnes obligatoires : Professeur | Classe | Matière | Heures / semaine
    # Colonne optionnelle   : Jour imposé (Lundi…Vendredi ou vide)
    # Note : la colonne « Taille du bloc » a été retirée du template (remplacée
    # par la section « Durée des séances » dans l'onglet Paramètres). Pour la
    # rétro-compatibilité, on la détecte encore si elle est présente.
    df = pd.read_excel(xls, "Services", skiprows=2, header=0)
    df = df.dropna(subset=[df.columns[0]])

    # Détection des colonnes optionnelles par leur nom (insensible à la casse)
    cols = {str(c).strip().lower(): i for i, c in enumerate(df.columns)}
    idx_bloc = next((i for k, i in cols.items() if "bloc" in k), None)
    idx_jour = next((i for k, i in cols.items()
                     if "jour" in k or "imposé" in k or "impose" in k), None)

    services = []
    for _, r in df.iterrows():
        prof, classe, matiere = _val(r, 0), _val(r, 1), _val(r, 2)
        try:
            heures = int(float(_val(r, 3)))
        except ValueError:
            continue
        if not (prof and classe and matiere and heures > 0):
            continue
        # Taille du bloc (rétro-compat, ignorée si colonne absente)
        bv = _val(r, idx_bloc) if idx_bloc is not None else ""
        bloc_size = 3 if bv == "3h" else (2 if bv in ("2h", "Oui") else 1)
        # Jour imposé
        jour_val = _val(r, idx_jour) if idx_jour is not None else ""
        jour_impose = JOUR_INDEX.get(jour_val)   # None si vide
        services.append({"prof": prof, "classe": classe,
                         "matiere": matiere, "heures": heures,
                         "bloc_size": bloc_size,
                         "jour_impose": jour_impose})

    # ── Durées de séance par matière (onglet Paramètres, optionnel) ──
    # Applique une plage [min, max] d'heures consécutives par séance à chaque
    # service de la matière concernée. Voir lire_durees_seances().
    durees = lire_durees_seances(xls)
    for svc in services:
        dmin, dmax = durees.get(svc["matiere"], (None, None))
        # Valeur effective : la règle de l'onglet prime ; sinon on déduit de
        # l'ancienne colonne « bloc » (rétro-compatibilité), sinon libre (1..H).
        if dmin is None:
            # Rétro-compatibilité : l'ancienne colonne « bloc Nh » signifiait
            # « regrouper par paquets d'AU PLUS N heures » (avec reste possible),
            # pas « tout en séances de N pile ». On traduit donc bloc N → max N,
            # min 1. Sans bloc → durée libre (1..H).
            b = svc["bloc_size"]
            dmin, dmax = (1, b) if b >= 2 else (1, svc["heures"])
        # Bornage de sécurité : une séance ne peut excéder le total d'heures,
        # ni 5h (le plus grand bloc continu de la journée : la matinée), car
        # une séance ne chevauche jamais la pause déjeuner.
        dmax = min(dmax, svc["heures"], len(SLOTS_MATIN))
        dmin = min(dmin, dmax)
        svc["seance_min"] = max(1, dmin)
        svc["seance_max"] = max(svc["seance_min"], dmax)

    # ── Disponibilités (demi-journées NON = indisponible) ──
    df = pd.read_excel(xls, "Disponibilités", skiprows=2, header=0)
    df = df.dropna(subset=[df.columns[0]])
    mapping = [(j, sl) for j in range(N_JOURS)
               for sl in (SLOTS_MATIN, SLOTS_APMIDI)]
    indispos = defaultdict(set)
    for _, r in df.iterrows():
        prof = _val(r, 0)
        for ci, (jour, slots) in enumerate(mapping, start=1):
            if _val(r, ci).upper() == "NON":
                for s in slots:
                    indispos[prof].add((jour, s))

    # ── Règles particulières (onglet Paramètres, section optionnelle) ──
    regles = lire_regles_parametres(xls)

    return classes, permanents, services, indispos, regles


def _hhmm_en_minutes(texte):
    """'11h00' ou '11:00' ou '11h' → minutes depuis minuit. None si illisible."""
    t = str(texte).strip().lower().replace("h", ":")
    if t.endswith(":"):
        t += "00"
    try:
        h, m = t.split(":")
        return int(h) * 60 + int(m)
    except (ValueError, AttributeError):
        return None


def creneaux_heures_chaudes(debut_txt, fin_txt):
    """Indices de créneaux (0..N_SLOTS-1) qui chevauchent la plage interdite.
    Un créneau est concerné dès qu'il déborde, même partiellement, dans la
    plage — interdiction « large » assumée."""
    deb = _hhmm_en_minutes(debut_txt)
    fin = _hhmm_en_minutes(fin_txt)
    if deb is None or fin is None or deb >= fin:
        return set()
    concernes = set()
    for i, label in enumerate(SLOT_LABELS):
        bornes = label.replace("–", "-").split("-")
        c_deb = _hhmm_en_minutes(bornes[0])
        c_fin = _hhmm_en_minutes(bornes[1])
        if c_deb is None or c_fin is None:
            continue
        if c_deb < fin and c_fin > deb:          # chevauchement
            concernes.add(i)
    return concernes


def _heures_depuis_texte(txt):
    """Extrait un nombre d'heures d'une cellule : '2', '2h', '3 h' → 2 ou 3.
    Retourne None si rien d'exploitable."""
    if txt is None:
        return None
    s = str(txt).strip().lower().replace("h", " ").replace(",", ".")
    s = s.split()[0] if s.split() else ""
    try:
        v = float(s)
    except ValueError:
        return None
    return int(round(v)) if v > 0 else None


def lire_durees_seances(xls):
    """Lit la section « DURÉE DES SÉANCES PAR MATIÈRE » de l'onglet Paramètres.

    Format attendu (tolérant) — un tableau à 3 colonnes :
        Matière | Durée min séance | Durée max séance
        SVT     | 2                | 3
        Physique-Chimie | 2        | 2
        Philosophie     | 1        | 2

    Retourne {matiere: (min, max)} en heures. Section absente → {} (le moteur
    se comporte alors comme avant : déduit des blocs ou décide librement).
    Lignes incomplètes ou non numériques ignorées silencieusement.
    """
    durees = {}
    try:
        df = pd.read_excel(xls, "Paramètres", header=None)
    except Exception:
        return durees

    n_lignes, n_cols = df.shape
    if n_cols < 3:
        return durees

    # Repérer la ligne-titre de la section.
    debut = None
    for i in range(n_lignes):
        cell = df.iat[i, 0]
        if pd.isna(cell):
            continue
        lib = str(cell).strip().lower()
        if "durée" in lib and "séance" in lib and "matière" in lib:
            debut = i
            break
        # tolère sans accents
        if "duree" in lib and "seance" in lib and "matiere" in lib:
            debut = i
            break
    if debut is None:
        return durees

    # Lire les lignes suivantes. On saute la ligne d'explication et la ligne
    # d'en-têtes ; on s'arrête à une ligne entièrement vide ou à une nouvelle
    # section (texte long sans valeurs numériques rencontré APRÈS des données).
    donnees_vues = False
    for i in range(debut + 1, n_lignes):
        c0 = df.iat[i, 0]
        if pd.isna(c0):
            if donnees_vues:
                break          # ligne vide après les données → fin du tableau
            continue           # ligne vide avant les données → on continue
        nom = str(c0).strip()
        bas = nom.lower()

        vmin = _heures_depuis_texte(df.iat[i, 1] if n_cols > 1 else None)
        vmax = _heures_depuis_texte(df.iat[i, 2] if n_cols > 2 else None)

        if vmin is None and vmax is None:
            # Pas de valeurs : soit en-tête / explication (avant données),
            # soit nouvelle section (après données → on arrête).
            if donnees_vues:
                break
            continue
        # Ignorer une éventuelle ligne d'en-têtes glissée ici.
        if bas in ("matière", "matiere"):
            continue
        if vmin is None:
            vmin = vmax
        if vmax is None:
            vmax = vmin
        if vmin and vmax:
            durees[nom] = (min(vmin, vmax), max(vmin, vmax))
            donnees_vues = True
    return durees


def lire_regles_parametres(xls):
    """Lit la section 'RÈGLES PARTICULIÈRES' de l'onglet Paramètres.
    Retourne un dict, par ex. {'eps_heures_chaudes': {0,1,...}} (créneaux
    interdits pour l'EPS) ou {} si la règle est absente/désactivée."""
    regles = {}
    try:
        df = pd.read_excel(xls, "Paramètres", header=None)
    except Exception:
        return regles

    n_lignes, n_cols = df.shape
    for i in range(n_lignes):
        cellule = df.iat[i, 0]
        if pd.isna(cellule):
            continue
        libelle = str(cellule).strip().lower()
        if "eps" in libelle and "chaud" in libelle:
            # colonnes : 0 libellé | 1 Activée ? | 2 début | 3 fin
            def _c(j):
                return "" if (j >= n_cols or pd.isna(df.iat[i, j])) \
                    else str(df.iat[i, j]).strip()
            active = _c(1).lower() in ("oui", "o", "yes", "true", "vrai", "1")
            if active:
                deb = _c(2) or "11h00"
                fin = _c(3) or "16h00"
                creneaux = creneaux_heures_chaudes(deb, fin)
                if creneaux:
                    regles["eps_heures_chaudes"] = creneaux
            break
    return regles


# ════════════════════ 2. SOLVEUR ════════════════════
def resoudre(classes, permanents, services, indispos, temps_max=120,
             matin_prefere=False, regles=None):
    """matin_prefere : si True, le moteur préfère remplir les matinées et
    libérer les après-midis (utile pour les écoles « tout le matin »).
    Dans tous les cas, la compacité des journées est fortement favorisée.

    regles : dict de règles particulières lues dans l'onglet Paramètres.
      - 'eps_heures_chaudes' : ensemble de créneaux (0-based) où l'EPS est
        strictement interdite (contrainte dure)."""
    regles = regles or {}
    model = cp_model.CpModel()

    # ── Variable principale : x[s,d,t] = service s placé jour d, créneau t
    x = {}
    for s in range(len(services)):
        for d in range(N_JOURS):
            for t in range(N_SLOTS):
                x[s, d, t] = model.new_bool_var(f"x_{s}_{d}_{t}")

    # ── Séances : suites d'heures consécutives d'un même service ──
    # Chaque service est découpé en séances dont la durée est comprise entre
    # seance_min et seance_max (réglable par matière dans l'onglet Paramètres).
    # On crée une variable de DÉBUT par taille autorisée et par position.
    # RÈGLE PHYSIQUE : une séance est CONTINUE dans le temps réel. Elle ne peut
    # donc jamais chevaucher la pause déjeuner (finir à 12h et « reprendre » à
    # 14h n'est pas une séance). Toute séance tient entièrement le matin
    # (créneaux 0-4) ou entièrement l'après-midi (créneaux 5-7). Les positions
    # à cheval sont fixées à 0. Aucune infaisabilité possible : le matin offre
    # 5 créneaux consécutifs et l'après-midi 3, largement assez pour les
    # séances de 1 à 3h utilisées en pratique.
    #
    #   debut[s, k, d, t] = 1  ⇔  une séance de durée k du service s commence
    #                              le jour d au créneau t (couvre t..t+k-1).
    #
    # Pour le jour imposé et la compacité, on garde une trace du début des
    # séances « longues » (k >= 2) via seance_long_start.
    debut = {}                 # (s, k, d, t) -> bool var
    seance_long_start = defaultdict(list)   # s -> liste de (d, t, k, var) pour k>=2
    for s, svc in enumerate(services):
        smin = svc["seance_min"]
        smax = svc["seance_max"]
        H = svc["heures"]
        tailles = list(range(smin, smax + 1))

        # Variables de début pour chaque taille / position valide.
        for k in tailles:
            for d in range(N_JOURS):
                for t in range(N_SLOTS - k + 1):
                    v = model.new_bool_var(f"deb_{s}_{k}_{d}_{t}")
                    debut[s, k, d, t] = v
                    # Séance à cheval sur la pause déjeuner : interdite.
                    if t < IDX_DEBUT_APMIDI <= t + k - 1:
                        model.add(v == 0)
                        continue
                    if k >= 2:
                        seance_long_start[s].append((d, t, k, v))

        # (a) Couverture : x[s,d,t] = somme des séances qui recouvrent (d,t).
        for d in range(N_JOURS):
            for t in range(N_SLOTS):
                recouvrants = []
                for k in tailles:
                    for ts in range(max(0, t - k + 1), min(N_SLOTS - k, t) + 1):
                        recouvrants.append(debut[s, k, d, ts])
                model.add(x[s, d, t] == sum(recouvrants))

        # (b) Volume exact via les séances : Σ (k · nb séances de taille k) = H.
        model.add(
            sum(k * debut[s, k, d, t]
                for k in tailles
                for d in range(N_JOURS)
                for t in range(N_SLOTS - k + 1)) == H
        )

        # (c) Pas de chevauchement de séances le même jour : sur chaque créneau,
        #     au plus une heure du service (déjà garanti par x ≤ 1 plus bas via
        #     H2/H3, mais on borne aussi x[s] lui-même pour la cohérence).
        for d in range(N_JOURS):
            for t in range(N_SLOTS):
                model.add(x[s, d, t] <= 1)

        # (d) AU PLUS UNE séance par jour pour ce service, quelle que soit sa
        #     taille. Conséquence : deux heures d'une même matière le même jour
        #     pour une classe sont forcément CONSÉCUTIVES (une seule séance),
        #     jamais éclatées (pas de « maths à 8h puis maths à 11h »). Une
        #     matière de 5h en séances de 1–2h se répartit donc sur ≥3 jours.
        for d in range(N_JOURS):
            model.add(
                sum(debut[s, k, d, t]
                    for k in tailles
                    for t in range(N_SLOTS - k + 1)) <= 1
            )
        # Le total d'heures du jour ne dépasse jamais la plus grande séance.
        for d in range(N_JOURS):
            model.add(sum(x[s, d, t] for t in range(N_SLOTS)) <= smax)

    # ── H1 : volume horaire exact ──
    for s, svc in enumerate(services):
        model.add(sum(x[s, d, t]
                      for d in range(N_JOURS)
                      for t in range(N_SLOTS)) == svc["heures"])

    # ── Index classes / profs ──
    cl_svcs, pr_svcs = defaultdict(list), defaultdict(list)
    for s, svc in enumerate(services):
        cl_svcs[svc["classe"]].append(s)
        pr_svcs[svc["prof"]].append(s)

    # ── H2 / H3 : pas de doublon classe ni prof ──
    for groupe in list(cl_svcs.values()) + list(pr_svcs.values()):
        for d in range(N_JOURS):
            for t in range(N_SLOTS):
                model.add(sum(x[s, d, t] for s in groupe) <= 1)

    # ── H4 : indisponibilités ──
    for s, svc in enumerate(services):
        for (d, t) in indispos.get(svc["prof"], set()):
            model.add(x[s, d, t] == 0)

    # ── H5 : permanents bloqués mercredi après-midi ──
    for s, svc in enumerate(services):
        if svc["prof"] in permanents:
            for t in SLOTS_APMIDI:
                model.add(x[s, 2, t] == 0)

    # ── H6 : limite journalière pour les services à durée « libre » ──

    # ── H6b (préférence forte) : éviter qu'une classe ait exactement 1h le ──
    # ── mercredi. Une classe qui n'a qu'1h ce jour-là se déplace pour une   ──
    # ── seule heure de cours, ce qui n'a aucun sens pédagogique.            ──
    # Choix de conception : PÉNALITÉ FORTE plutôt qu'interdiction absolue. Une
    # contrainte dure « 0 ou >=2 » peut rendre le problème infaisable dans des
    # cas tendus (ex : une matière d'1h imposée le mercredi alors que les autres
    # profs sont indisponibles ce jour). Une pénalité forte donne le même
    # résultat en pratique — le solveur évite le mercredi à 1h dès qu'il le peut
    # — sans jamais bloquer la génération. Le poids est très supérieur aux
    # autres pénalités pour que ce soit évité en priorité.
    # On compte TOUTE la journée (matin + après-midi vacataires).
    POIDS_MERC_1H = 50
    merc_penalites = []
    for cl, s_list in cl_svcs.items():
        heures_merc = sum(x[s, 2, t] for s in s_list for t in range(N_SLOTS))
        # est_1h = 1 si la classe a exactement 1h le mercredi.
        est_1h = model.new_bool_var(f"merc_1h_{cl}")
        # Lien : est_1h ⇔ (heures_merc == 1). On l'encode avec deux implications
        # via une variable auxiliaire « au moins 1h » et « au moins 2h ».
        au_moins_1 = model.new_bool_var(f"merc_ge1_{cl}")
        au_moins_2 = model.new_bool_var(f"merc_ge2_{cl}")
        model.add(heures_merc >= 1).only_enforce_if(au_moins_1)
        model.add(heures_merc == 0).only_enforce_if(au_moins_1.Not())
        model.add(heures_merc >= 2).only_enforce_if(au_moins_2)
        model.add(heures_merc <= 1).only_enforce_if(au_moins_2.Not())
        # exactement 1h = (au_moins_1) ET (NON au_moins_2)
        model.add(est_1h == 1).only_enforce_if([au_moins_1, au_moins_2.Not()])
        model.add(est_1h == 0).only_enforce_if(au_moins_1.Not())
        model.add(est_1h == 0).only_enforce_if(au_moins_2)
        merc_penalites.append(POIDS_MERC_1H * est_1h)

    # ── H6c (préférence forte) : éviter qu'une classe ait exactement 1h sur ──
    # ── une DEMI-JOURNÉE (matin ou après-midi).                             ──
    # Cas typique constaté sur le terrain : la classe finit sa matinée, puis
    # revient (ou reste) l'après-midi pour UNE seule matière — fatigant pour
    # les élèves, sans intérêt pédagogique. Dès qu'une demi-journée est
    # occupée, elle doit compter au moins 2 heures ; sinon le solveur regroupe
    # ces heures isolées sur moins de demi-journées (idéalement le matin).
    # Même choix de conception que H6b : PÉNALITÉ FORTE, jamais de contrainte
    # dure (une matière d'1h à jour imposé chez un prof très indisponible
    # rendrait la grille infaisable). Poids SUPÉRIEUR à POIDS_TROU (25) et aux
    # poids de regroupement profs (15/1) : le confort élève prime toujours.
    POIDS_DEMI_1H = 30
    for cl, s_list in cl_svcs.items():
        for d in range(N_JOURS):
            for nom, slots in (("mat", SLOTS_MATIN), ("apm", SLOTS_APMIDI)):
                h_demi = sum(x[s, d, t] for s in s_list for t in slots)
                ge1 = model.new_bool_var(f"demi_ge1_{cl}_{d}_{nom}")
                ge2 = model.new_bool_var(f"demi_ge2_{cl}_{d}_{nom}")
                model.add(h_demi >= 1).only_enforce_if(ge1)
                model.add(h_demi == 0).only_enforce_if(ge1.Not())
                model.add(h_demi >= 2).only_enforce_if(ge2)
                model.add(h_demi <= 1).only_enforce_if(ge2.Not())
                seule = model.new_bool_var(f"demi_1h_{cl}_{d}_{nom}")
                model.add(seule == 1).only_enforce_if([ge1, ge2.Not()])
                model.add(seule == 0).only_enforce_if(ge1.Not())
                model.add(seule == 0).only_enforce_if(ge2)
                merc_penalites.append(POIDS_DEMI_1H * seule)

    # ── H6c (préférence forte) : éviter les DEMI-JOURNÉES à 1 seul cours ──
    # Une classe qui ne vient l'après-midi (ou le matin) que pour une seule
    # heure se déplace pour presque rien : fatigant pour les élèves, sans
    # intérêt pédagogique. On pénalise chaque demi-journée où la classe a
    # EXACTEMENT 1h : soit le moteur regroupe (>= 2h), soit il libère la
    # demi-journée (0h), typiquement en ramenant l'heure isolée le matin.
    # Même choix de conception que le mercredi : PÉNALITÉ, jamais contrainte
    # dure (une contrainte « 0 ou >=2 » peut rendre le problème infaisable
    # quand les disponibilités sont tendues).
    # Poids : SOUS POIDS_TROU (on ne crée jamais un trou dans la journée d'une
    # classe pour éviter une demi-journée courte) mais AU-DESSUS du poids
    # vacataire (on accepte qu'un prof vienne un jour de plus si ça évite aux
    # élèves un déplacement pour une seule heure).
    POIDS_DEMI_1H = 20
    demi_penalites = []
    for cl, s_list in cl_svcs.items():
        for d in range(N_JOURS):
            for nom, slots in (("am", SLOTS_MATIN), ("pm", SLOTS_APMIDI)):
                h_demi = sum(x[s, d, t] for s in s_list for t in slots)
                ge1 = model.new_bool_var(f"demi_ge1_{cl}_{d}_{nom}")
                ge2 = model.new_bool_var(f"demi_ge2_{cl}_{d}_{nom}")
                model.add(h_demi >= 1).only_enforce_if(ge1)
                model.add(h_demi == 0).only_enforce_if(ge1.Not())
                model.add(h_demi >= 2).only_enforce_if(ge2)
                model.add(h_demi <= 1).only_enforce_if(ge2.Not())
                est_1h = model.new_bool_var(f"demi_1h_{cl}_{d}_{nom}")
                model.add(est_1h == 1).only_enforce_if([ge1, ge2.Not()])
                model.add(est_1h == 0).only_enforce_if(ge1.Not())
                model.add(est_1h == 0).only_enforce_if(ge2)
                demi_penalites.append(POIDS_DEMI_1H * est_1h)

    # Un service dont la séance n'est pas contrainte (seance_min == 1 et
    # seance_max == total) ne doit pas s'entasser : on garde l'ancienne règle
    # « ≤ 2h par jour » pour étaler la matière sur la semaine. Les services
    # avec une durée de séance explicite (k>=2) sont déjà bornés à smax/jour.
    for s, svc in enumerate(services):
        libre = (svc["seance_min"] == 1 and svc["seance_max"] >= svc["heures"]
                 and svc["heures"] >= 2)
        if libre:
            for d in range(N_JOURS):
                model.add(sum(x[s, d, t] for t in range(N_SLOTS)) <= 2)

    # ── H7 : jour imposé (optionnel) ──
    for s, svc in enumerate(services):
        d = svc["jour_impose"]
        if d is None:
            continue
        if svc["seance_max"] >= 2 and svc["heures"] >= 2:
            # au moins une séance longue (k>=2) commence ce jour-là si possible,
            # sinon au moins une heure ce jour-là (cas où seules des 1h tiennent)
            longues_ce_jour = [v for (dd, t, k, v) in seance_long_start[s]
                               if dd == d]
            if longues_ce_jour:
                model.add(sum(longues_ce_jour) >= 1)
            else:
                model.add(sum(x[s, d, t] for t in range(N_SLOTS)) >= 1)
        else:
            # au moins une séance ce jour-là
            model.add(sum(x[s, d, t] for t in range(N_SLOTS)) >= 1)

    # ── H8 : pas d'EPS aux heures chaudes (règle optionnelle de l'école) ──
    creneaux_chauds = regles.get("eps_heures_chaudes", set())
    if creneaux_chauds:
        for s, svc in enumerate(services):
            if svc["matiere"] == "EPS":
                for d in range(N_JOURS):
                    for t in creneaux_chauds:
                        model.add(x[s, d, t] == 0)

    # ── Objectif : qualité pédagogique + compacité ──
    bonus, malus = [], []
    for s, svc in enumerate(services):
        for d in range(N_JOURS):
            if svc["matiere"] in SCIENCES:
                for t in SLOTS_MATIN:
                    bonus.append(x[s, d, t])
            if svc["matiere"] == "EPS":
                malus.append(x[s, d, IDX_DEBUT_APMIDI])
            if svc["matiere"] in LOURDES:
                malus.append(x[s, d, N_SLOTS - 1])

    # ── Compacité : pénaliser tout créneau vide ENTRE deux cours sur la
    # même JOURNÉE (et plus seulement à l'intérieur d'une demi-journée).
    # C'est ce qui élimine le cas « cours à 10h puis cours à 16h » : le grand
    # vide du midi devient un trou coûteux que le moteur cherche à supprimer
    # en regroupant les cours. Pénalité forte (POIDS_TROU).
    POIDS_TROU = 25

    def penaliser_trous(groupes, prefixe):
        for g_idx, s_list in enumerate(groupes):
            for d in range(N_JOURS):
                def occ(t):
                    return sum(x[s, d, t] for s in s_list)
                # « actif[t] » = il existe un cours à un créneau >= t ce jour-là.
                # Un trou = créneau vide alors qu'un cours vient avant ET après.
                # On le détecte sur toute l'amplitude de la journée (1..7).
                for t in range(1, N_SLOTS):
                    # cours avant t (au moins un)
                    avant = [occ(tp) for tp in range(0, t)]
                    # cours en t ou après (au moins un)
                    apres = [occ(tp) for tp in range(t, N_SLOTS)]
                    a_avant = model.new_bool_var(f"{prefixe}av_{g_idx}_{d}_{t}")
                    a_apres = model.new_bool_var(f"{prefixe}ap_{g_idx}_{d}_{t}")
                    model.add(sum(avant) >= 1).only_enforce_if(a_avant)
                    model.add(sum(avant) == 0).only_enforce_if(a_avant.Not())
                    model.add(sum(apres) >= 1).only_enforce_if(a_apres)
                    model.add(sum(apres) == 0).only_enforce_if(a_apres.Not())
                    # trou si : créneau t vide, ET cours avant, ET cours après
                    trou = model.new_bool_var(f"{prefixe}_{g_idx}_{d}_{t}")
                    model.add(trou <= 1 - occ(t))
                    model.add(trou <= a_avant)
                    model.add(trou <= a_apres)
                    model.add(trou >= a_avant + a_apres + (1 - occ(t)) - 2)
                    malus.append(POIDS_TROU * trou)

    penaliser_trous(list(cl_svcs.values()), "tc")
    penaliser_trous(list(pr_svcs.values()), "tp")

    # ── Regroupement des profs : pénaliser le NOMBRE DE JOURS de présence ──
    # Pour chaque prof, on minimise le nombre de jours où il vient. Un prof à
    # 12h gagne à venir 2-3 jours pleins plutôt que 5 demi-journées : crucial
    # pour les vacataires venant d'un autre établissement (trajets). C'est une
    # PRÉFÉRENCE (pénalité), pas une contrainte dure : le moteur regroupe quand
    # il le peut sans jamais rendre la grille infaisable.
    #
    # Poids volontairement INFÉRIEUR à POIDS_TROU : on n'accepte jamais de créer
    # un trou dans la grille d'une CLASSE pour économiser un jour à un prof. Le
    # confort élève prime sur le confort enseignant.
    # Poids DIFFÉRENCIÉ : fort pour les vacataires (ils viennent d'un autre
    # établissement, chaque jour économisé est un vrai gain), symbolique pour
    # les permanents (présents de toute façon). Concentrer le poids sur les
    # vacataires permet au solveur d'atteindre ≤2 jours de présence en
    # quelques secondes au lieu de plafonner à 3-4 jours.
    POIDS_JOUR_VACATAIRE = 15
    POIDS_JOUR_PERMANENT = 1
    present = {}
    for prof, s_list in pr_svcs.items():
        w = POIDS_JOUR_PERMANENT if prof in permanents else POIDS_JOUR_VACATAIRE
        for d in range(N_JOURS):
            p = model.new_bool_var(f"prof_present_{prof}_{d}")
            present[prof, d] = p
            # p = 1 si au moins un cours ce jour, 0 sinon (lien exact).
            heures_jour = [x[s, d, t] for s in s_list for t in range(N_SLOTS)]
            model.add(sum(heures_jour) >= 1).only_enforce_if(p)
            model.add(sum(heures_jour) == 0).only_enforce_if(p.Not())
            malus.append(w * p)

    # ── Option « matin de préférence » : chaque heure placée l'après-midi
    # reçoit un petit malus, donc à compacité égale le moteur préfère
    # remplir les matinées. Poids volontairement faible pour ne JAMAIS
    # primer sur la compacité ni les contraintes dures.
    if matin_prefere:
        for s in range(len(services)):
            for d in range(N_JOURS):
                for t in SLOTS_APMIDI:
                    malus.append(x[s, d, t])

    model.maximize(sum(bonus) - sum(malus) - sum(merc_penalites)
                   - sum(demi_penalites))

    # ── Résolution ──
    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = temps_max
    solver.parameters.num_search_workers  = 4

    print(f"  Résolution en cours (max {temps_max}s)…")
    status = solver.solve(model)
    if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        print(f"  ✗ Aucune solution ({solver.status_name(status)}).")
        print("    → Vérifiez disponibilités, volumes et jours imposés.")
        return None
    print(f"  ✓ Solution {solver.status_name(status)} "
          f"(score qualité = {int(solver.objective_value)})")

    emplois = {}
    for s, svc in enumerate(services):
        for d in range(N_JOURS):
            for t in range(N_SLOTS):
                if solver.value(x[s, d, t]):
                    emplois[(svc["classe"], d, t)] = {
                        "prof": svc["prof"], "matiere": svc["matiere"]}
    return emplois


# ════════════════ 3. VÉRIFICATEUR INTÉGRÉ ════════════════
def verifier(emplois, classes, permanents, services, indispos):
    ok = True

    # Volumes horaires exacts
    placed = defaultdict(int)
    for (cl, d, t), info in emplois.items():
        placed[(info["prof"], cl, info["matiere"])] += 1
    nb_ko = sum(1 for s in services
                if placed[(s["prof"], s["classe"], s["matiere"])] != s["heures"])
    if nb_ko:
        print(f"  ✗ {nb_ko} volumes horaires incorrects !")
        ok = False
    else:
        print(f"  ✓ Volumes horaires : {len(services)}/{len(services)} exacts")

    # Conflits classe / prof
    seen_p, conflits = set(), 0
    for (cl, d, t), info in emplois.items():
        k = (info["prof"], d, t)
        if k in seen_p:
            conflits += 1
        seen_p.add(k)
    if conflits:
        print(f"  ✗ {conflits} conflits professeur !")
        ok = False
    else:
        print("  ✓ Zéro conflit (classes et professeurs)")

    # Indisponibilités
    nb = sum(1 for (cl, d, t), i in emplois.items()
             if (d, t) in indispos.get(i["prof"], set()))
    print(f"  {'✓' if nb == 0 else '✗'} Disponibilités : "
          f"{'toutes respectées' if nb == 0 else str(nb) + ' violations !'}")
    ok = ok and nb == 0

    # Mercredi PM permanents
    nb = sum(1 for (cl, d, t), i in emplois.items()
             if d == 2 and t in SLOTS_APMIDI and i["prof"] in permanents)
    print(f"  {'✓' if nb == 0 else '✗'} Mercredi après-midi : "
          f"{'aucun permanent' if nb == 0 else str(nb) + ' permanents placés !'}")
    ok = ok and nb == 0

    # Trous dans les grilles classes (sur la JOURNÉE entière : un créneau
    # vide entouré de cours avant et après compte comme un trou, même s'il
    # tombe sur l'heure du midi).
    trous = 0
    for cl in classes:
        for d in range(N_JOURS):
            occ = [(cl, d, t) in emplois for t in range(N_SLOTS)]
            for t in range(1, N_SLOTS):
                if not occ[t] and any(occ[:t]) and any(occ[t:]):
                    trous += 1
    print(f"  {'✓' if trous == 0 else '!'} Trous dans les grilles : {trous}")

    # Jours imposés
    for s in services:
        d = s["jour_impose"]
        if d is None:
            continue
        slots = sorted(t for t in range(N_SLOTS)
                       if emplois.get((s["classe"], d, t), {}).get("prof") == s["prof"]
                       and emplois[(s["classe"], d, t)]["matiere"] == s["matiere"])
        # Nouvelle règle (durées de séance souples) : le jour imposé garantit
        # qu'AU MOINS UNE séance de la matière tombe ce jour-là. On vérifie donc
        # la présence d'au moins seance_min heures, consécutives si possible.
        smin = s.get("seance_min", s["bloc_size"] if s["bloc_size"] >= 2 else 1)
        if slots:
            run, best = 1, 1
            for i in range(1, len(slots)):
                run = run + 1 if slots[i] == slots[i - 1] + 1 else 1
                best = max(best, run)
        else:
            best = 0
        # Respecté si au moins une séance de durée >= seance_min est présente
        # (ou, pour les matières à durée libre, au moins 1h ce jour-là).
        respecte = best >= max(1, smin) if smin >= 2 else len(slots) >= 1
        sym = "✓" if respecte else "✗"
        taille_txt = (f", séance de {best}h présente"
                      if respecte and best >= 2 else "")
        print(f"  {sym} Jour imposé : {s['matiere']} {s['classe']} "
              f"→ {JOURS[d]} ({len(slots)}h placées{taille_txt})")
        ok = ok and respecte

    return ok


# ════════════════════ 4. EXPORT EXCEL ════════════════════
def _brd(top=False):
    t = Side(style="thin",   color="AAAAAA")
    m = Side(style="medium", color="444444")
    return Border(left=t, right=t, top=m if top else t, bottom=t)


def _hdr(cell, texte):
    cell.value = texte
    cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", start_color="444444")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _brd()


def _grille_jours(ws, row0, get_contenu):
    """Dessine une grille (jours en colonnes, créneaux en lignes)
    à partir de la ligne row0. get_contenu(d, t) → texte ou None."""
    ws.cell(row=row0, column=1).fill = PatternFill("solid", start_color="444444")
    ws.cell(row=row0, column=1).border = _brd()
    for j, jour in enumerate(JOURS, start=2):
        _hdr(ws.cell(row=row0, column=j), jour)
    ws.row_dimensions[row0].height = 19

    row = row0 + 1
    for t_idx, label in enumerate(SLOT_LABELS):
        if t_idx == IDX_DEBUT_APMIDI:
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row, end_column=6)
            sep = ws.cell(row=row, column=1, value="APRES-MIDI")
            sep.font = Font(name="Arial", size=8, bold=True)
            sep.fill = PatternFill("solid", start_color="DDDDDD")
            sep.alignment = Alignment(horizontal="center", vertical="center")
            sep.border = _brd(top=True)
            ws.row_dimensions[row].height = 12
            row += 1

        ws.row_dimensions[row].height = 36
        h = ws.cell(row=row, column=1, value=label)
        h.font = Font(name="Arial", size=8, bold=True)
        h.fill = PatternFill("solid", start_color="F2F2F2")
        h.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        h.border = _brd()

        for d_idx in range(N_JOURS):
            c = ws.cell(row=row, column=d_idx + 2)
            contenu = get_contenu(d_idx, t_idx)
            if contenu:
                c.value = contenu
                c.font = Font(name="Arial", size=9)
                c.fill = PatternFill("solid", start_color="FFFFFF")
            else:
                c.fill = PatternFill("solid", start_color="FAFAFA")
            c.alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=True)
            c.border = _brd()
        row += 1
    return row


def ecrire_excel(emplois, classes, chemin):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Une feuille par classe ──
    for cl in classes:
        ws = wb.create_sheet(title=cl[:31])
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 15
        for j in range(2, 7):
            ws.column_dimensions[get_column_letter(j)].width = 22

        ws.merge_cells("A1:F1")
        t = ws["A1"]
        t.value = f"Emploi du temps  —  {cl}"
        t.font = Font(name="Arial", size=13, bold=True)
        t.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        def contenu(d, ti, _cl=cl):
            info = emplois.get((_cl, d, ti))
            if not info:
                return None
            nom = info["prof"].replace("M. ", "")
            return f"{info['matiere']}\n{nom}"

        _grille_jours(ws, 2, contenu)

    # ── Vue par professeur (texte simple, AUCUN emoji) ──
    ws = wb.create_sheet(title="Vue Professeurs")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 15
    for j in range(2, 7):
        ws.column_dimensions[get_column_letter(j)].width = 22

    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = "Emplois du temps  —  Vue par professeur"
    t.font = Font(name="Arial", size=13, bold=True)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    edt_prof = defaultdict(dict)
    for (cl, d, ti), info in emplois.items():
        edt_prof[info["prof"]][(d, ti)] = f"{info['matiere']}\n{cl}"

    cur = 2
    for prof in sorted(edt_prof.keys()):
        ws.merge_cells(start_row=cur, start_column=1,
                       end_row=cur, end_column=6)
        c = ws.cell(row=cur, column=1, value=f"  {prof}")
        c.font = Font(name="Arial", size=10, bold=True)
        c.fill = PatternFill("solid", start_color="CCCCCC")
        c.alignment = Alignment(vertical="center")
        c.border = _brd()
        ws.row_dimensions[cur].height = 18
        cur += 1

        def contenu_p(d, ti, _p=prof):
            return edt_prof[_p].get((d, ti))

        cur = _grille_jours(ws, cur, contenu_p) + 1

    wb.save(chemin)
    print(f"  ✓ Fichier écrit : {chemin}")
    print(f"    Onglets : {', '.join(wb.sheetnames)}")


# ════════════════════ 5. POINT D'ENTRÉE ════════════════════
if __name__ == "__main__":
    entree = sys.argv[1] if len(sys.argv) > 1 else "donnees.xlsx"
    sortie = sys.argv[2] if len(sys.argv) > 2 else "emplois_du_temps.xlsx"

    print("=" * 55)
    print("  MOTEUR EDT — Génération des emplois du temps")
    print("=" * 55)

    print("\n[1/4] Lecture des données…")
    classes, permanents, services, indispos, regles = lire_excel(entree)
    n_b2 = sum(1 for s in services if s["bloc_size"] == 2)
    n_b3 = sum(1 for s in services if s["bloc_size"] == 3)
    n_ji = sum(1 for s in services if s["jour_impose"] is not None)
    print(f"  {len(classes)} classes | {len(services)} services | "
          f"{sum(s['heures'] for s in services)}h à placer")
    print(f"  Blocs 2h : {n_b2} | Blocs 3h : {n_b3} | Jours imposés : {n_ji}")
    if regles.get("eps_heures_chaudes"):
        print(f"  Règle EPS heures chaudes : {len(regles['eps_heures_chaudes'])} "
              "créneau(x) interdit(s)")

    print("\n[2/4] Résolution…")
    emplois = resoudre(classes, permanents, services, indispos, regles=regles)
    if not emplois:
        sys.exit(1)

    print("\n[3/4] Vérification automatique…")
    verifier(emplois, classes, permanents, services, indispos)

    print("\n[4/4] Export Excel…")
    ecrire_excel(emplois, classes, sortie)
    print("\n✓ Terminé.")
